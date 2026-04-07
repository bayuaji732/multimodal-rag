"""
ingestion/parser.py
───────────────────
Extracts text, images, and tables from uploaded files.

Supported formats & table understanding:
  ┌─────────────┬────────────────────────────────────────────────┐
  │ Format      │ Table support                                  │
  ├─────────────┼────────────────────────────────────────────────┤
  │ PDF         │ ✓ pdfplumber bounding-box extraction           │
  │ CSV         │ ✓ native — entire file is one/many tables      │
  │ XLSX / XLS  │ ✓ every sheet → separate table chunk(s)       │
  │ DOCX        │ ✓ unstructured table elements                  │
  │ PNG/JPG/…   │ ✗ image chunk (no tabular extraction)          │
  └─────────────┴────────────────────────────────────────────────┘

Large CSV / XLSX sheets are paginated into TABLE_PAGE_SIZE-row chunks
so embeddings stay within token limits and retrieval stays precise.
"""
from __future__ import annotations

import base64
import io
import logging
import re
import uuid
from pathlib import Path

import pdfplumber
from PIL import Image
from unstructured.partition.auto import partition
from unstructured.documents.elements import (
    Table, Image as UnstructuredImage, Text, Title, NarrativeText,
)

from utils.models import ChunkType, DocumentChunk

log = logging.getLogger(__name__)

TABLE_PAGE_SIZE = 50   # max data rows per table chunk for CSV/XLSX


# ─── shared helpers ───────────────────────────────────────────────────────────

def _pil_to_b64(img: Image.Image, max_px: int = 1024) -> str:
    img.thumbnail((max_px, max_px), Image.LANCZOS)
    buf = io.BytesIO()
    img.save(buf, format="PNG")
    return base64.b64encode(buf.getvalue()).decode()


def _clean_text(t: str) -> str:
    return re.sub(r"\s+", " ", t).strip()


# ─── table normalization ──────────────────────────────────────────────────────

def _normalize_table(raw: list[list]) -> list[list[str]]:
    """Stringify, clean, drop empty rows and columns."""
    if not raw:
        return []
    rows: list[list[str]] = [
        [_clean_text(str(cell)) if cell is not None else "" for cell in row]
        for row in raw
    ]
    rows = [r for r in rows if any(c for c in r)]
    if not rows:
        return []
    max_cols = max(len(r) for r in rows)
    rows = [r + [""] * (max_cols - len(r)) for r in rows]
    non_empty_cols = [
        col for col in range(max_cols)
        if any(rows[row][col] for row in range(len(rows)))
    ]
    return [[r[c] for c in non_empty_cols] for r in rows]


def _is_header_row(row: list[str]) -> bool:
    if not row:
        return False
    non_empty = [c for c in row if c]
    if not non_empty:
        return False
    numeric = sum(1 for c in non_empty if re.match(r"^[\d,.\-\s%$€£]+$", c))
    title   = sum(1 for c in non_empty if c == c.title() or c == c.upper())
    return numeric == 0 and title / len(non_empty) >= 0.6


def _table_to_markdown(rows: list[list[str]], title: str | None = None) -> str:
    if not rows:
        return ""
    parts: list[str] = []
    if title:
        parts.append(f"**{title}**\n")
    header_row = rows[0]
    data_rows  = rows[1:]
    if not _is_header_row(header_row):
        data_rows  = rows
        header_row = [f"Column {i+1}" for i in range(len(rows[0]))]
    esc = lambda c: c.replace("|", "\\|")
    parts.append("| " + " | ".join(esc(c) for c in header_row) + " |")
    parts.append("| " + " | ".join("---" for _ in header_row) + " |")
    for row in data_rows:
        parts.append("| " + " | ".join(esc(c) for c in row) + " |")
    return "\n".join(parts)


def _make_table_chunk(
    rows: list[list[str]],
    doc_id: str,
    doc_name: str,
    page: int,
    title: str | None = None,
    bbox: tuple | None = None,
) -> DocumentChunk | None:
    if len(rows) < 2:
        return None
    header_row = rows[0] if _is_header_row(rows[0]) else []
    data_rows  = rows[1:] if header_row else rows
    md = _table_to_markdown(rows, title=title)
    if not md:
        return None
    return DocumentChunk(
        doc_id        = doc_id,
        doc_name      = doc_name,
        chunk_type    = ChunkType.TABLE,
        text          = md,
        page          = page,
        bbox          = tuple(bbox) if bbox else None,
        table_title   = title,
        table_headers = header_row,
        table_rows    = len(data_rows),
        table_cols    = len(rows[0]),
    )


def _paginate_table(
    header: list[str],
    data: list[list[str]],
    doc_id: str,
    doc_name: str,
    base_title: str,
    base_page: int,
) -> list[DocumentChunk]:
    """Split a large table into TABLE_PAGE_SIZE-row chunks."""
    chunks: list[DocumentChunk] = []
    total_pages = max(1, (len(data) + TABLE_PAGE_SIZE - 1) // TABLE_PAGE_SIZE)
    for pn in range(total_pages):
        start    = pn * TABLE_PAGE_SIZE
        end      = start + TABLE_PAGE_SIZE
        slice_   = data[start:end]
        ptitle   = base_title
        if total_pages > 1:
            ptitle = f"{base_title} (rows {start+1}–{min(end, len(data))})"
        chunk = _make_table_chunk(
            [header] + slice_,
            doc_id, doc_name,
            page=base_page + pn,
            title=ptitle,
        )
        if chunk:
            chunks.append(chunk)
    return chunks


# ─── PDF parser ───────────────────────────────────────────────────────────────

def _detect_table_title(page, table_bbox) -> str | None:
    x0, y0, x1, _ = table_bbox
    try:
        words = page.within_bbox(
            (x0 - 20, max(0, y0 - 50), x1 + 20, y0)
        ).extract_words()
    except Exception:
        return None
    if not words:
        return None
    line = " ".join(w["text"] for w in words).strip()
    if re.match(r"(?i)^(table|figure|tab\.?|fig\.?|exhibit)\b", line) or len(line) < 80:
        return line or None
    return None


class PDFParser:
    def parse(self, path: Path, doc_id: str, doc_name: str | None = None) -> list[DocumentChunk]:
        chunks: list[DocumentChunk] = []
        doc_name = doc_name or path.name

        with pdfplumber.open(path) as pdf:
            for page_num, page in enumerate(pdf.pages, start=1):
                tables_with_bbox = page.find_tables()
                for tbl_obj in tables_with_bbox:
                    raw = tbl_obj.extract()
                    normalized = _normalize_table(raw or [])
                    title = _detect_table_title(page, tbl_obj.bbox)
                    chunk = _make_table_chunk(
                        normalized, doc_id, doc_name,
                        page=page_num, title=title, bbox=tbl_obj.bbox,
                    )
                    if chunk:
                        chunks.append(chunk)

                non_table_page = page
                for tbl_obj in tables_with_bbox:
                    try:
                        non_table_page = non_table_page.outside_bbox(tbl_obj.bbox)
                    except Exception:
                        pass
                text = non_table_page.extract_text(x_tolerance=2, y_tolerance=2) or ""
                text = _clean_text(text)
                if len(text) > 40:
                    chunks.append(DocumentChunk(
                        doc_id=doc_id, doc_name=doc_name,
                        chunk_type=ChunkType.TEXT, text=text, page=page_num,
                    ))

        try:
            elements = partition(
                filename=str(path), strategy="hi_res",
                extract_images_in_pdf=True,
            )
            for el in elements:
                if isinstance(el, UnstructuredImage):
                    p    = getattr(el.metadata, "page_number", 0) or 0
                    ip   = getattr(el.metadata, "image_path", None)
                    b64  = ""
                    if ip and Path(ip).exists():
                        b64 = _pil_to_b64(Image.open(ip).convert("RGB"))
                    if b64:
                        chunks.append(DocumentChunk(
                            doc_id=doc_id, doc_name=doc_name,
                            chunk_type=ChunkType.IMAGE,
                            text=_clean_text(str(el)) or f"Figure on page {p}",
                            image_b64=b64, page=p,
                        ))
        except Exception as exc:
            log.warning("unstructured image pass failed: %s", exc)

        log.info("PDFParser: %d chunks from %s", len(chunks), path.name)
        return chunks


# ─── CSV parser ───────────────────────────────────────────────────────────────

class CSVParser:
    """
    Entire CSV = table(s). Paginated into TABLE_PAGE_SIZE-row chunks.
    Encoding detection: utf-8-sig → utf-8 → latin-1.
    Delimiter detection: csv.Sniffer on first 4 KB.
    """

    def parse(self, path: Path, doc_id: str, doc_name: str | None = None) -> list[DocumentChunk]:
        import csv

        doc_name  = doc_name or path.name
        raw_rows: list[list[str]] = []

        for enc in ("utf-8-sig", "utf-8", "latin-1"):
            try:
                with path.open(encoding=enc, newline="") as f:
                    sample = f.read(4096)
                    f.seek(0)
                    try:
                        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
                    except csv.Error:
                        dialect = csv.excel
                    raw_rows = list(csv.reader(f, dialect))
                break
            except (UnicodeDecodeError, Exception):
                continue

        if not raw_rows:
            log.warning("CSVParser: could not read %s", path.name)
            return []

        normalized = _normalize_table(raw_rows)
        if len(normalized) < 2:
            log.warning("CSVParser: %s has < 2 rows after normalization", path.name)
            return []

        header = normalized[0]
        data   = normalized[1:]
        title  = f"CSV: {path.stem}"

        chunks = _paginate_table(header, data, doc_id, doc_name, title, base_page=1)
        log.info("CSVParser: %d chunk(s) from %s (%d data rows)", len(chunks), path.name, len(data))
        return chunks


# ─── XLSX / XLS parser ────────────────────────────────────────────────────────

class XLSXParser:
    """
    Each worksheet → one or more TABLE chunks (paginated).
    Uses openpyxl (data_only=True so formulas show their computed values).
    """

    def parse(self, path: Path, doc_id: str, doc_name: str | None = None) -> list[DocumentChunk]:
        import openpyxl

        doc_name = doc_name or path.name
        chunks:  list[DocumentChunk] = []

        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        except Exception as exc:
            log.error("XLSXParser: cannot open %s — %s", path.name, exc)
            return []

        for sheet_idx, sheet_name in enumerate(wb.sheetnames, start=1):
            ws = wb[sheet_name]
            raw_rows: list[list[str]] = [
                [str(cell) if cell is not None else "" for cell in row]
                for row in ws.iter_rows(values_only=True)
            ]
            normalized = _normalize_table(raw_rows)
            if len(normalized) < 2:
                log.debug("Sheet '%s': < 2 useful rows, skipping.", sheet_name)
                continue

            header = normalized[0]
            data   = normalized[1:]
            title  = f"{path.stem} — {sheet_name}"
            # Use sheet_idx * 1000 as base page so chunks from different sheets
            # don't collide on page numbers
            sheet_chunks = _paginate_table(
                header, data, doc_id, doc_name, title,
                base_page=sheet_idx * 1000,
            )
            chunks.extend(sheet_chunks)

        wb.close()
        log.info("XLSXParser: %d chunk(s) from %s", len(chunks), path.name)
        return chunks


# ─── Image parser ─────────────────────────────────────────────────────────────

class ImageParser:
    def parse(self, path: Path, doc_id: str, doc_name: str | None = None) -> list[DocumentChunk]:
        img = Image.open(path).convert("RGB")
        b64 = _pil_to_b64(img)
        return [DocumentChunk(
            doc_id=doc_id, doc_name=doc_name or path.name,
            chunk_type=ChunkType.IMAGE,
            text=f"Image: {path.stem}", image_b64=b64, page=1,
        )]


# ─── DOCX parser ──────────────────────────────────────────────────────────────

class DocxParser:
    def parse(self, path: Path, doc_id: str, doc_name: str | None = None) -> list[DocumentChunk]:
        elements  = partition(filename=str(path))
        chunks:   list[DocumentChunk] = []
        _doc_name = doc_name or path.name
        for el in elements:
            if isinstance(el, (Text, Title, NarrativeText)):
                t = _clean_text(str(el))
                if len(t) > 20:
                    chunks.append(DocumentChunk(
                        doc_id=doc_id, doc_name=_doc_name,
                        chunk_type=ChunkType.TEXT, text=t,
                    ))
            elif isinstance(el, Table):
                raw_text = _clean_text(el.text)
                if raw_text:
                    chunks.append(DocumentChunk(
                        doc_id=doc_id, doc_name=_doc_name,
                        chunk_type=ChunkType.TABLE, text=raw_text,
                    ))
        return chunks


# ─── dispatcher ───────────────────────────────────────────────────────────────

SUFFIX_MAP: dict[str, type] = {
    ".pdf":  PDFParser,
    ".csv":  CSVParser,
    ".xlsx": XLSXParser,
    ".xls":  XLSXParser,
    ".png":  ImageParser,
    ".jpg":  ImageParser,
    ".jpeg": ImageParser,
    ".webp": ImageParser,
    ".docx": DocxParser,
}

SUPPORTED_FORMATS = sorted(SUFFIX_MAP.keys())


def parse_document(
    path: Path,
    doc_id: str | None = None,
    doc_name: str | None = None,
) -> list[DocumentChunk]:
    doc_id     = doc_id or str(uuid.uuid4())
    suffix     = path.suffix.lower()
    parser_cls = SUFFIX_MAP.get(suffix)
    if parser_cls is None:
        raise ValueError(
            f"Unsupported file type: '{suffix}'. "
            f"Supported: {', '.join(SUPPORTED_FORMATS)}"
        )
    return parser_cls().parse(path, doc_id, doc_name=doc_name)


# ─── text chunker ─────────────────────────────────────────────────────────────

def chunk_text(
    chunks: list[DocumentChunk],
    chunk_size: int = 512,
    overlap: int = 64,
) -> list[DocumentChunk]:
    """
    Split TEXT chunks exceeding chunk_size.
    TABLE and IMAGE chunks pass through unchanged — tables must never be split
    mid-row (pagination happens at parse time for CSV/XLSX).
    """
    result: list[DocumentChunk] = []
    for chunk in chunks:
        if chunk.chunk_type != ChunkType.TEXT or len(chunk.text.split()) <= chunk_size:
            result.append(chunk)
            continue
        words = chunk.text.split()
        start = 0
        while start < len(words):
            end = min(start + chunk_size, len(words))
            result.append(DocumentChunk(
                doc_id=chunk.doc_id, doc_name=chunk.doc_name,
                chunk_type=chunk.chunk_type,
                text=" ".join(words[start:end]),
                page=chunk.page,
            ))
            start += chunk_size - overlap
    return result