"""
streamlit_app.py  —  Multi-Modal RAG Knowledge Engine  v3
──────────────────────────────────────────────────────────
Lightweight rewrite — same visual design, half the code.

New in v3:
  • Orchestration toggle (LangGraph multi-hop decomposition)
  • RAGAS evaluation option (faithfulness / relevancy / precision)
  • Trace ID displayed per query
  • Cleaner CSS via consolidated helper functions
"""
from __future__ import annotations

import io
from pathlib import Path

import httpx
import streamlit as st

API_BASE = "http://localhost:8000"

# ─── Page config ──────────────────────────────────────────────────────────────

st.set_page_config(
    page_title="RAG · Knowledge Engine",
    page_icon="◈",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ─── HTTP helpers ─────────────────────────────────────────────────────────────

@st.cache_resource
def _client() -> httpx.Client:
    return httpx.Client(
        base_url=API_BASE,
        timeout=httpx.Timeout(connect=3, read=90, write=30, pool=5),
        limits=httpx.Limits(max_connections=10, max_keepalive_connections=5),
    )


def api(method: str, path: str, **kw):
    try:
        r = _client().request(method, path, **kw)
        return (r.json(), None) if r.status_code < 400 else (None, f"HTTP {r.status_code}: {r.text[:200]}")
    except httpx.ConnectError:
        return None, "Cannot connect to API at `localhost:8000`."
    except Exception as e:
        return None, str(e)


@st.cache_data(ttl=15, show_spinner=False)
def _health() -> bool:
    _, err = api("GET", "/health")
    return err is None


@st.cache_data(ttl=30, show_spinner=False)
def _docs() -> list:
    data, _ = api("GET", "/documents")
    return data or []


# ─── HTML micro-helpers ───────────────────────────────────────────────────────

def _tag(text: str, cls: str = "td") -> str:
    return f'<span class="tag {cls}">{text}</span>'


def _badge(text: str, cls: str = "b-ok") -> str:
    return f'<span class="badge {cls}">{text}</span>'


_FMT = {
    ".pdf":  ("PDF",  "tc"), ".csv":  ("CSV",  "tg"), ".xlsx": ("XLSX", "tg"),
    ".xls":  ("XLS",  "tg"), ".docx": ("DOCX", "to"), ".png":  ("PNG",  "tc"),
    ".jpg":  ("JPG",  "tc"), ".jpeg": ("JPEG", "tc"), ".webp": ("WEBP", "tc"),
}
UPLOAD_TYPES = [e.lstrip(".") for e in _FMT]


def _fmt_tag(name: str) -> str:
    label, cls = _FMT.get(Path(name).suffix.lower(), ("?", "td"))
    return _tag(label, cls)


def _html(content: str):
    st.markdown(content, unsafe_allow_html=True)


def _section(label: str):
    _html(f'<div class="sl">{label}</div>')


# ─── CSS (single injection) ──────────────────────────────────────────────────

@st.cache_data(show_spinner=False)
def _css() -> str:
    return """<style>
@import url('https://fonts.googleapis.com/css2?family=Syne:wght@700;800&family=IBM+Plex+Mono:wght@300;400;500&display=swap');
html,body,[class*="css"]{font-family:'IBM Plex Mono',monospace!important;background:#080b10;color:#e8edf3}
#MainMenu,footer,header{visibility:hidden}
.block-container{padding:2rem 2.5rem 4rem;max-width:1260px}
[data-testid="stSidebar"]{background:#0a0f18;border-right:1px solid #1a2535}
[data-testid="stSidebar"] *{font-family:'IBM Plex Mono',monospace!important}
.pt{font-family:'Syne',sans-serif!important;font-size:2.4rem;font-weight:800;letter-spacing:-.02em;line-height:1.1;color:#e8edf3;margin-bottom:.2rem}
.pt span{color:#00e5ff}
.ps{font-size:.68rem;color:#5a6a7a;letter-spacing:.1em;text-transform:uppercase;margin-bottom:1.8rem}
.sl{font-size:.58rem;letter-spacing:.22em;text-transform:uppercase;color:#5a6a7a;margin-bottom:.7rem;display:flex;align-items:center;gap:8px}
.sl::after{content:'';flex:1;height:1px;background:#1a2535}
.tag{display:inline-block;font-size:.58rem;letter-spacing:.1em;text-transform:uppercase;padding:3px 9px;border-radius:2px;margin:2px;white-space:nowrap}
.tc{background:rgba(0,229,255,.08);color:#00e5ff;border:1px solid rgba(0,229,255,.25)}
.to{background:rgba(255,107,53,.08);color:#ff6b35;border:1px solid rgba(255,107,53,.25)}
.tg{background:rgba(127,255,107,.08);color:#7fff6b;border:1px solid rgba(127,255,107,.25)}
.td{background:rgba(255,255,255,.04);color:#5a6a7a;border:1px solid #1a2535}
.tp{background:rgba(168,85,247,.08);color:#c084fc;border:1px solid rgba(168,85,247,.25)}
.card{background:#0d1520;border:1px solid #1a2535;border-radius:3px;padding:1.25rem;margin-bottom:.75rem}
.ans{background:#090e18;border:1px solid #1a2535;border-top:2px solid #00e5ff;padding:1.4rem;border-radius:3px;font-size:.85rem;line-height:1.85;color:#e8edf3;white-space:pre-wrap}
.uv{background:rgba(255,107,53,.12);color:#ff6b35;border-radius:2px;padding:1px 5px}
.ans-md table{border-collapse:collapse;width:100%;margin:.8rem 0;font-size:.78rem}
.ans-md th{background:rgba(0,229,255,.06);color:#00e5ff;padding:.45rem .7rem;border:1px solid #1a2535;text-align:left;font-weight:500}
.ans-md td{padding:.4rem .7rem;border:1px solid #1a2535;color:#9ab0c5}
.ans-md tr:nth-child(even) td{background:rgba(255,255,255,.015)}
.cit{background:#090e18;border-left:3px solid #00e5ff;padding:.7rem 1rem;margin:.35rem 0;border-radius:0 3px 3px 0;font-size:.74rem}
.cit-table{border-left-color:#7fff6b} .cit-image{border-left-color:#ff6b35}
.cit-m{font-size:.58rem;letter-spacing:.1em;text-transform:uppercase;color:#5a6a7a;margin-bottom:.35rem}
.cit-t{color:#9ab0c5;line-height:1.6;margin-top:.3rem}
.cit-headers{display:flex;flex-wrap:wrap;gap:4px;margin-top:.45rem}
.ch{font-size:.55rem;padding:2px 7px;background:rgba(127,255,107,.06);color:#7fff6b;border:1px solid rgba(127,255,107,.2);border-radius:2px}
.warn{background:rgba(255,107,53,.06);border:1px solid rgba(255,107,53,.3);border-radius:3px;padding:.7rem 1rem;font-size:.72rem;color:#ff6b35;margin-top:.5rem}
.sg{display:flex;gap:.75rem;margin-bottom:1.4rem;flex-wrap:wrap}
.sc{flex:1;min-width:110px;background:#0d1520;border:1px solid #1a2535;border-radius:3px;padding:.9rem;text-align:center}
.sv{font-family:'Syne',sans-serif;font-size:1.9rem;font-weight:700;color:#00e5ff;line-height:1}
.sk{font-size:.58rem;letter-spacing:.15em;text-transform:uppercase;color:#5a6a7a;margin-top:3px}
.badge{display:inline-block;padding:3px 10px;border-radius:2px;font-size:.58rem;letter-spacing:.12em;text-transform:uppercase}
.b-ok{background:rgba(127,255,107,.1);color:#7fff6b;border:1px solid rgba(127,255,107,.3)}
.b-pend{background:rgba(0,229,255,.1);color:#00e5ff;border:1px solid rgba(0,229,255,.3)}
.b-err{background:rgba(255,107,53,.1);color:#ff6b35;border:1px solid rgba(255,107,53,.3)}
.tr{display:flex;gap:1rem;font-size:.62rem;color:#5a6a7a;letter-spacing:.08em;text-transform:uppercase;margin-top:.5rem;flex-wrap:wrap}
.tv{color:#9ab0c5}
.dr{display:flex;align-items:center;justify-content:space-between;background:#090e18;border:1px solid #1a2535;border-radius:3px;padding:.65rem 1.1rem;margin-bottom:.4rem;font-size:.75rem}
.dn{color:#e8edf3;font-weight:600} .dm{color:#5a6a7a;font-size:.6rem;margin-top:2px}
.tbl-preview{overflow-x:auto;border:1px solid #1a2535;border-radius:3px;background:#090e18;padding:.6rem}
.tbl-preview table{border-collapse:collapse;font-size:.7rem;width:100%;min-width:400px}
.tbl-preview th{background:rgba(127,255,107,.06);color:#7fff6b;padding:.35rem .6rem;border:1px solid #1a2535;font-weight:500;white-space:nowrap}
.tbl-preview td{padding:.3rem .6rem;border:1px solid #1a2535;color:#9ab0c5}
.tbl-preview tr:nth-child(even) td{background:rgba(255,255,255,.015)}
.ragas-bar{height:6px;border-radius:3px;margin-top:2px}
div[data-testid="stTextInput"] input,div[data-testid="stTextArea"] textarea{background:#090e18!important;border:1px solid #1a2535!important;color:#e8edf3!important;font-family:'IBM Plex Mono',monospace!important;border-radius:3px!important}
div[data-testid="stTextInput"] input:focus,div[data-testid="stTextArea"] textarea:focus{border-color:#00e5ff!important;box-shadow:0 0 0 1px rgba(0,229,255,.2)!important}
div[data-testid="stFileUploader"]{background:#090e18;border:1px dashed #1a2535!important;border-radius:3px}
.stButton>button{background:transparent!important;border:1px solid #00e5ff!important;color:#00e5ff!important;font-family:'IBM Plex Mono',monospace!important;font-size:.7rem!important;letter-spacing:.08em;border-radius:2px!important;padding:.45rem 1.1rem!important;transition:background .15s,color .15s}
.stButton>button:hover{background:#00e5ff!important;color:#080b10!important}
.stCheckbox label{font-size:.72rem!important;color:#9ab0c5!important}
hr{border:none;border-top:1px solid #1a2535;margin:1.2rem 0}
</style>"""


_html(_css())

# ─── Session state ────────────────────────────────────────────────────────────

for k, v in {"page": "query", "last_result": None, "ingest_jobs": []}.items():
    st.session_state.setdefault(k, v)


# ─── Sidebar ──────────────────────────────────────────────────────────────────

with st.sidebar:
    _html(
        '<div style="padding:1.2rem 0 .8rem">'
        '<div style="font-size:.56rem;letter-spacing:.22em;text-transform:uppercase;color:#5a6a7a;margin-bottom:7px">⬡  LLM / GenAI Engineer</div>'
        '<div style="font-family:\'Syne\',sans-serif;font-size:1.35rem;font-weight:800;color:#e8edf3;line-height:1.1">'
        'Multi-Modal<br/><span style="color:#00e5ff">RAG Engine</span>'
        '</div></div><hr/>'
    )
    _html(
        _badge("● API Online", "b-ok") if _health() else _badge("● API Offline", "b-err")
    )
    _html("<br/>")

    for key, label in [("query", "◈  Query"), ("ingest", "↑  Ingest"), ("documents", "≡  Documents")]:
        if st.button(label, key=f"nav_{key}", use_container_width=True):
            if st.session_state.page != key:
                st.session_state.page = key
                st.rerun()

    _html("<hr/>")
    _section("Settings")
    apply_guard    = st.checkbox("NLI Hallucination Guard", value=True)
    use_stream     = st.checkbox("Streaming Response", value=False)
    use_orchestr   = st.checkbox("Multi-hop Orchestration", value=True)
    run_ragas      = st.checkbox("RAGAS Evaluation", value=False)

    _html("<hr/>")
    _section("Supported Formats")
    legend = [
        (".pdf", "tc", "Text · Tables · Images"), (".csv", "tg", "Table (full file)"),
        (".xlsx", "tg", "Table (per sheet)"), (".docx", "to", "Text · Tables"),
        (".png", "tc", "Image (CLIP)"), (".jpg", "tc", "Image (CLIP)"),
    ]
    _html(
        '<div style="margin-top:.3rem">'
        + "".join(
            f'<div style="display:flex;align-items:center;gap:10px;padding:.45rem 0;'
            f'border-bottom:1px solid #0f1825;font-size:.68rem">'
            f'{_tag(ext, cls)}'
            f'<span style="color:#5a6a7a;font-size:.62rem">{desc}</span></div>'
            for ext, cls, desc in legend
        )
        + "</div>"
    )
    _html(
        '<br/><div style="font-size:.58rem;color:#5a6a7a;line-height:1.8">'
        "RAG · Hybrid BM25+Dense<br>CLIP · text-embedding-3<br>"
        "Qdrant · Cohere Rerank<br>GPT-4o · NLI Guard</div>"
    )


# ══════════════════════════════════════════════════════════════════════════════
# RENDER HELPERS
# ══════════════════════════════════════════════════════════════════════════════

def _render_table_preview(uploaded) -> None:
    """HTML preview of the first 10 rows of a CSV / XLSX."""
    ext = Path(uploaded.name).suffix.lower()
    try:
        if ext == ".csv":
            import csv as _csv, io as _io
            raw = uploaded.getvalue()
            for enc in ("utf-8-sig", "utf-8", "latin-1"):
                try:
                    text = raw.decode(enc)
                    try:
                        dialect = _csv.Sniffer().sniff(text[:4096], delimiters=",;\t|")
                    except _csv.Error:
                        dialect = _csv.excel
                    rows = [r for r in _csv.reader(_io.StringIO(text), dialect) if any(c.strip() for c in r)]
                    break
                except UnicodeDecodeError:
                    continue
        elif ext in (".xlsx", ".xls"):
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(uploaded.getvalue()), read_only=True, data_only=True)
            ws = wb[wb.sheetnames[0]]
            rows = [
                [str(c) if c is not None else "" for c in row]
                for row in ws.iter_rows(values_only=True)
            ]
            rows = [r for r in rows if any(c.strip() for c in r)]
            wb.close()
        else:
            return

        if not rows:
            return
        header, data = rows[0], rows[1:11]
        total = len(rows) - 1
        ths = "".join(f"<th>{h}</th>" for h in header)
        trs = "".join("<tr>" + "".join(f"<td>{c}</td>" for c in r) + "</tr>" for r in data)
        more = f'<div style="font-size:.62rem;color:#5a6a7a;padding:.5rem;text-align:center;border-top:1px solid #1a2535">… {total - 10:,} more rows</div>' if total > 10 else ""
        _html(f'<div class="tbl-preview"><table><thead><tr>{ths}</tr></thead><tbody>{trs}</tbody></table>{more}</div>')
        _html(f'<div style="font-size:.6rem;color:#5a6a7a;margin-top:.35rem">{len(header)} columns · {total:,} rows</div>')
    except Exception as exc:
        _html(f'<div class="warn">Preview unavailable: {exc}</div>')


def _render_ragas(scores: dict | None) -> None:
    """Render RAGAS quality scores as mini bar charts."""
    if not scores:
        return
    _section("Quality Scores (RAGAS)")
    cols = st.columns(len(scores))
    palette = {"faithfulness": "#7fff6b", "answer_relevancy": "#00e5ff", "context_precision": "#c084fc"}
    for col, (metric, val) in zip(cols, scores.items()):
        with col:
            color = palette.get(metric, "#9ab0c5")
            pct = max(0, min(100, val * 100))
            label = metric.replace("_", " ").title()
            _html(
                f'<div class="card" style="padding:.7rem;text-align:center">'
                f'<div style="font-size:1.3rem;font-weight:700;color:{color};font-family:Syne,sans-serif">{pct:.0f}%</div>'
                f'<div style="background:#1a2535;border-radius:3px;overflow:hidden;margin-top:4px">'
                f'<div class="ragas-bar" style="width:{pct}%;background:{color}"></div></div>'
                f'<div style="font-size:.52rem;letter-spacing:.12em;text-transform:uppercase;color:#5a6a7a;margin-top:5px">{label}</div>'
                f'</div>'
            )


def _render_result(data: dict) -> None:
    """Render answer + citations + optional RAGAS."""
    answer = data.get("answer", "")
    citations = data.get("citations", [])
    warnings = data.get("warnings", [])

    _section("Answer")
    display = answer.replace("[⚠ unverified]", '<span class="uv">[⚠ unverified]</span>')

    if "|" in answer and "---" in answer:
        _html(f'<div class="ans-md" style="background:#090e18;border:1px solid #1a2535;border-top:2px solid #00e5ff;padding:1.4rem;border-radius:3px;font-size:.85rem;line-height:1.85">')
        st.markdown(display, unsafe_allow_html=True)
        _html("</div>")
    else:
        _html(f'<div class="ans">{display}</div>')

    # Metadata row: model · tokens · trace
    trace_id = data.get("trace_id", "")
    trace_chip = f' <span>Trace: <span class="tv">{trace_id[:12]}…</span></span>' if trace_id else ""
    _html(
        f'<div class="tr">'
        f'<span>Model: <span class="tv">{data.get("model","")}</span></span>'
        f'<span>Prompt: <span class="tv">{data.get("prompt_tokens",0):,}</span> tok</span>'
        f'<span>Completion: <span class="tv">{data.get("completion_tokens",0):,}</span> tok</span>'
        f'{trace_chip}</div>'
    )

    if warnings:
        _html(
            f'<div class="warn">⚠ NLI Guard flagged {len(warnings)} sentence(s):<br>'
            + "<br>".join(f"• {w}" for w in warnings)
            + "</div>"
        )

    # RAGAS scores
    _render_ragas(data.get("ragas_scores"))

    if not citations:
        return

    _html('<div style="margin-top:1.1rem">')
    _section("Sources & Citations")
    parts: list[str] = []
    for c in citations:
        ct = c.get("chunk_type", "text")
        type_tag = _tag(ct, {"text": "td", "image": "tc", "table": "tg"}.get(ct, "td"))
        img_b = _tag("📷 img", "tc") if c.get("has_image") else ""

        # Table metadata
        tmeta = ""
        if ct == "table":
            r, co = c.get("table_rows"), c.get("table_cols")
            dims = _tag(f"{r}r × {co}c", "tg") if r is not None and co is not None else ""
            title = c.get("table_title") or ""
            title_line = f'<div style="font-size:.65rem;color:#7fff6b;margin:.3rem 0 .1rem">↳ {title}</div>' if title else ""
            hdrs = c.get("table_headers") or []
            chips = "".join(f'<span class="ch">{h}</span>' for h in hdrs[:8])
            chips += f'<span class="ch" style="opacity:.5">+{len(hdrs)-8}</span>' if len(hdrs) > 8 else ""
            h_html = f'<div class="cit-headers">{chips}</div>' if chips else ""
            tmeta = title_line + (f'<div style="display:flex;align-items:center;gap:6px;margin-top:.2rem">{dims}{h_html}</div>' if dims or h_html else "")

        parts.append(
            f'<div class="cit cit-{ct}">'
            f'<div class="cit-m">[{c["index"]}] &nbsp; {c["doc_name"]} &nbsp;·&nbsp; p.{c["page"]} &nbsp;{type_tag} {img_b}</div>'
            f'{tmeta}<div class="cit-t">{c.get("text_snippet","")}</div></div>'
        )
    parts.append("</div>")
    _html("".join(parts))


# ══════════════════════════════════════════════════════════════════════════════
# PAGE: QUERY
# ══════════════════════════════════════════════════════════════════════════════

def page_query():
    _html(
        '<div class="pt">Ask the<br/><span>Knowledge Base</span></div>'
        '<div class="ps">Multi-modal retrieval · Grounded citations · NLI verified</div>'
    )
    docs = _docs()
    col_q, col_f = st.columns([3, 1])
    with col_q:
        query = st.text_area(
            "q", placeholder="e.g. Which region had the highest revenue in Q3?",
            height=120, label_visibility="collapsed", key="query_input",
        )
    with col_f:
        _section("Filter Docs")
        opts = {f"{d['doc_name']} ({d['chunk_count']}ch)": d["doc_id"] for d in docs}
        selected = st.multiselect("Docs", list(opts), label_visibility="collapsed")
        filter_ids = [opts[l] for l in selected] or None

    if st.button("◈  Run Query", disabled=not (query or "").strip(), key="run_query"):
        if use_stream:
            _run_streaming(query, filter_ids)
        else:
            endpoint = "/query/evaluate" if run_ragas else "/query"
            with st.spinner("Retrieving & generating…"):
                data, err = api("POST", endpoint, json={
                    "query": query,
                    "apply_guard": apply_guard,
                    "filter_doc_ids": filter_ids,
                    "use_orchestration": use_orchestr,
                })
            if err:
                _html(f'<div class="warn">⚠ {err}</div>')
            else:
                st.session_state.last_result = data
                _render_result(data)
    elif st.session_state.last_result:
        _render_result(st.session_state.last_result)


def _run_streaming(query: str, filter_ids):
    _section("Answer")
    ph = st.empty()
    full = ""
    params: dict = {"q": query}
    if filter_ids:
        params["doc_ids"] = ",".join(filter_ids)
    try:
        with httpx.stream("GET", f"{API_BASE}/query/stream", params=params, timeout=120) as r:
            for line in r.iter_lines():
                if line.startswith("data: "):
                    tok = line[6:]
                    if tok == "[DONE]":
                        break
                    full += tok
                    ph.markdown(f'<div class="ans">{full}▌</div>', unsafe_allow_html=True)
        ph.markdown(f'<div class="ans">{full}</div>', unsafe_allow_html=True)
    except Exception as e:
        st.error(f"Streaming error: {e}")


# ══════════════════════════════════════════════════════════════════════════════
# PAGE: INGEST
# ══════════════════════════════════════════════════════════════════════════════

def page_ingest():
    _html(
        '<div class="pt">Ingest<br/><span>Documents</span></div>'
        '<div class="ps">PDF · CSV · XLSX · DOCX · PNG · JPG · Async Celery pipeline</div>'
    )
    col_up, col_jobs = st.columns(2, gap="large")

    with col_up:
        _section("Upload File")
        uploaded = st.file_uploader("Drop file", type=UPLOAD_TYPES, label_visibility="collapsed")

        if uploaded:
            ext = Path(uploaded.name).suffix.lower()
            _html(
                f'<div class="card" style="border-top:2px solid #7fff6b;padding:.9rem 1.1rem;margin-top:.5rem">'
                f'<div style="display:flex;justify-content:space-between;align-items:center">'
                f'<div><div style="font-size:.6rem;color:#5a6a7a;text-transform:uppercase">Ready to ingest</div>'
                f'<div style="font-size:.88rem;font-weight:600;color:#e8edf3;margin-top:2px">{uploaded.name}</div>'
                f'<div style="font-size:.6rem;color:#5a6a7a">{uploaded.size/1024:.1f} KB</div></div>'
                f'{_fmt_tag(uploaded.name)}</div></div>'
            )
            if ext in (".csv", ".xlsx", ".xls"):
                _section("Table Preview")
                _render_table_preview(uploaded)

            if st.button("↑  Start Ingestion", key="do_ingest"):
                with st.spinner("Uploading…"):
                    data, err = api("POST", "/ingest", files={"file": (uploaded.name, uploaded.getvalue(), uploaded.type)})
                if err:
                    _html(f'<div class="warn">⚠ {err}</div>')
                else:
                    st.session_state.ingest_jobs.insert(0, {
                        "task_id": data["task_id"], "doc_id": data["doc_id"],
                        "filename": data["filename"], "status": "PENDING",
                    })
                    _docs.clear()
                    st.toast(f"Queued: {data['filename']}", icon="📄")
                    st.rerun()

    with col_jobs:
        _job_panel()


@st.fragment(run_every=5)
def _job_panel():
    _section("Ingestion Jobs")
    jobs = st.session_state.ingest_jobs
    if not jobs:
        _html('<div style="font-size:.75rem;color:#5a6a7a;padding:.75rem 0">No jobs yet.</div>')
        return

    badge_map = {
        "SUCCESS": _badge("✓ done", "b-ok"), "FAILURE": _badge("✗ failed", "b-err"),
        "STARTED": _badge("⟳ processing", "b-pend"), "PROGRESS": _badge("⟳ processing", "b-pend"),
        "PENDING": _badge("… queued", "b-pend"),
    }
    for job in jobs:
        if job["status"] not in ("SUCCESS", "FAILURE"):
            data, _ = api("GET", f"/jobs/{job['task_id']}")
            if data and data["status"] != job["status"]:
                job["status"] = data["status"]
                if data.get("result"):
                    job["chunks"] = data["result"].get("chunk_count", "?")
                if data["status"] == "SUCCESS":
                    _docs.clear()

    parts = []
    for job in jobs:
        b = badge_map.get(job["status"], badge_map["PENDING"])
        ch = f'· {job.get("chunks","?")} chunks' if job["status"] == "SUCCESS" else ""
        parts.append(
            f'<div class="card" style="padding:.8rem 1rem;margin-bottom:.4rem">'
            f'<div style="display:flex;justify-content:space-between;align-items:center">'
            f'<span style="font-size:.78rem;font-weight:600;color:#e8edf3">{job["filename"]}</span>{b}</div>'
            f'<div style="display:flex;align-items:center;gap:8px;margin-top:.35rem">'
            f'{_fmt_tag(job["filename"])}'
            f'<span style="font-size:.58rem;color:#5a6a7a">{job["task_id"][:22]}… {ch}</span>'
            f'</div></div>'
        )
    _html("".join(parts))


# ══════════════════════════════════════════════════════════════════════════════
# PAGE: DOCUMENTS
# ══════════════════════════════════════════════════════════════════════════════

def page_documents():
    _html(
        '<div class="pt">Indexed<br/><span>Documents</span></div>'
        '<div class="ps">Manage the knowledge base · Qdrant collection</div>'
    )
    col_r, _ = st.columns([1, 5])
    with col_r:
        if st.button("↻  Refresh", key="refresh_docs"):
            _docs.clear()
            st.rerun()

    docs = _docs()
    total = sum(d.get("chunk_count", 0) for d in docs)
    avg = round(total / max(len(docs), 1))

    _html(
        f'<div class="sg">'
        f'<div class="sc"><div class="sv">{len(docs)}</div><div class="sk">Documents</div></div>'
        f'<div class="sc"><div class="sv">{total:,}</div><div class="sk">Total Chunks</div></div>'
        f'<div class="sc"><div class="sv">{avg}</div><div class="sk">Avg / Doc</div></div>'
        f'</div>'
    )

    if not docs:
        _html('<div style="font-size:.8rem;color:#5a6a7a;padding:.75rem 0">No documents indexed yet.</div>')
        return

    _section("Documents")
    for doc in docs:
        col_info, col_del = st.columns([6, 1])
        with col_info:
            _html(
                f'<div class="dr"><div>'
                f'<div class="dn">{doc["doc_name"]}</div>'
                f'<div class="dm">{doc["doc_id"][:28]}… &nbsp;·&nbsp; {doc["chunk_count"]} chunks</div>'
                f'</div>{_fmt_tag(doc["doc_name"])}</div>'
            )
        with col_del:
            if st.button("✕", key=f"del_{doc['doc_id']}", help="Delete"):
                _, err = api("DELETE", f"/documents/{doc['doc_id']}")
                if err:
                    st.error(err)
                else:
                    _docs.clear()
                    st.toast(f"Deleted {doc['doc_name']}", icon="🗑")
                    st.rerun()


# ─── Router ───────────────────────────────────────────────────────────────────

{"query": page_query, "ingest": page_ingest, "documents": page_documents}[st.session_state.page]()